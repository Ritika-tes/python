from selenium import webdriver
import openpyxl

filename = "D:\selenium excel\selenium.xlsx"
workbook = openpyxl.load_workbook("D:\selenium excel\selenium.xlsx")
sheet = workbook['sheet_1']

##### print number of rows and columns
row = sheet.max_row
column = sheet.max_column
print(row)
print(column)




#####  read cell values from sheet
print(sheet.cell(row=1,column=2).value)
print(sheet.cell(row=2,column=1).value)



######### write the cell values in sheet

sheet.cell(row = 6,column = 1).value = "5"
sheet.cell(row = 6,column = 2).value = "Delhi"
workbook.save("D:\selenium excel\selenium.xlsx")


################ print all values for rows of column excel file

for i in range (1,row+1):
    for j in range(1,column+1):
        print(sheet.cell(row=i,column=j).value,end=' ')
    print('\n')
